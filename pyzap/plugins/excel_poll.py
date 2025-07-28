"""Excel polling trigger."""

from __future__ import annotations

import os
import json
from typing import Any, Dict, List

from ..core import BaseTrigger


__all__ = [
    "ExcelPollTrigger",
    "ExcelCellTrigger",
    "ExcelFileTrigger",
    "ExcelAttachmentsTrigger",
]


class ExcelPollTrigger(BaseTrigger):
    """Poll an Excel workbook for newly appended rows."""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.state_file = self.config.get("state_file")
        self.start_row = int(self.config.get("start_row", 2))
        self.last_row = self.start_row - 1
        if self.state_file and os.path.exists(self.state_file):
            try:
                with open(self.state_file, "r", encoding="utf-8") as fh:
                    self.last_row = int(fh.read().strip() or self.last_row)
            except Exception:
                pass

    def _save_state(self) -> None:
        if not self.state_file:
            return
        try:
            with open(self.state_file, "w", encoding="utf-8") as fh:
                fh.write(str(self.last_row))
        except Exception:
            pass

    def poll(self) -> List[Dict[str, Any]]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pragma: no cover - dependency missing
            raise RuntimeError(
                "excel_poll trigger requires the 'openpyxl' package. Install it with 'pip install openpyxl'."
            ) from exc

        file_path = self.config.get("file")
        sheet_name = self.config.get("sheet")
        filters: Dict[Any, Any] = self.config.get("filters", {})
        if not file_path:
            raise ValueError("file parameter required")

        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        max_row = getattr(ws, "max_row", None)
        if max_row is None:
            rows_attr = getattr(ws, "rows", [])
            try:
                max_row = len(list(rows_attr))
            except TypeError:
                max_row = len(rows_attr)

        results: List[Dict[str, Any]] = []
        for row_idx in range(self.last_row + 1, max_row + 1):
            cells = ws[row_idx]
            values = [getattr(c, "value", None) for c in cells]
            match = True
            for col, expected in filters.items():
                idx = int(col) - 1
                val = values[idx] if idx < len(values) else None
                if val != expected:
                    match = False
                    break
            if match:
                results.append({"id": str(row_idx), "values": values})

        self.last_row = max_row
        self._save_state()
        return results


class ExcelAttachmentsTrigger(ExcelPollTrigger):
    """Extend :class:`ExcelPollTrigger` parsing an attachments column."""

    def poll(self) -> List[Dict[str, Any]]:
        rows = super().poll()
        col = int(self.config.get("attachments_col", 1)) - 1
        for row in rows:
            raw = row["values"][col] if col < len(row["values"]) else ""
            attachments: List[str] = []
            if isinstance(raw, str):
                attachments = [a.strip() for a in raw.split(",") if a.strip()]
            row["attachments"] = attachments
        return rows


class ExcelCellTrigger(BaseTrigger):
    """Trigger when values in selected columns change."""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.state_file = self.config.get("state_file")
        cols = self.config.get("columns", [])
        self.columns = [int(c) for c in cols]
        self.state: Dict[str, Dict[str, Any]] = {}
        if self.state_file and os.path.exists(self.state_file):
            try:
                with open(self.state_file, "r", encoding="utf-8") as fh:
                    self.state = json.load(fh)
            except Exception:
                self.state = {}
            self._initialized = True
        else:
            self._initialized = False

    def _save_state(self) -> None:
        if not self.state_file:
            return
        try:
            with open(self.state_file, "w", encoding="utf-8") as fh:
                json.dump(self.state, fh)
        except Exception:
            pass

    def poll(self) -> List[Dict[str, Any]]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pragma: no cover - dependency missing
            raise RuntimeError(
                "excel_cell trigger requires the 'openpyxl' package. Install it with 'pip install openpyxl'."
            ) from exc

        file_path = self.config.get("file")
        sheet_name = self.config.get("sheet")
        if not file_path:
            raise ValueError("file parameter required")

        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        max_row = getattr(ws, "max_row", len(getattr(ws, "rows", [])))
        results: List[Dict[str, Any]] = []
        for row_idx in range(1, max_row + 1):
            cells = ws[row_idx]
            values = [getattr(c, "value", None) for c in cells]
            row_state = self.state.get(str(row_idx), {})
            changed: Dict[str, Any] = {}
            for col in self.columns:
                idx = col - 1
                val = values[idx] if idx < len(values) else None
                if row_state.get(str(col)) != val:
                    changed[str(col)] = val
                    row_state[str(col)] = val
            if self._initialized and changed:
                results.append({"id": f"{row_idx}", "values": values, "changes": changed})
            if row_state:
                self.state[str(row_idx)] = row_state

        self._save_state()
        self._initialized = True
        return results


class ExcelFileTrigger(BaseTrigger):
    """Trigger when the Excel file modification time changes."""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.last_mtime: float | None = None

    def poll(self) -> List[Dict[str, Any]]:
        file_path = self.config.get("file")
        if not file_path:
            raise ValueError("file parameter required")
        if not os.path.exists(file_path):
            return []
        mtime = os.path.getmtime(file_path)
        if self.last_mtime is None:
            self.last_mtime = mtime
            return []
        if mtime > self.last_mtime:
            self.last_mtime = mtime
            return [{"id": str(int(mtime)), "file": file_path}]
        return []
