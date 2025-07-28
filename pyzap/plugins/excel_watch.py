"""Additional Excel triggers and actions."""

from __future__ import annotations

import json
import logging
import os
import re
from typing import Any, Dict, List

from ..core import BaseTrigger, BaseAction


class ExcelRowAddedTrigger(BaseTrigger):
    """Detect new rows with optional advanced filtering."""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.state_file = self.config.get("state_file")
        self.start_row = int(self.config.get("start_row", 2))
        self.last_row = self.start_row - 1
        if self.state_file and os.path.exists(self.state_file):
            try:
                with open(self.state_file, "r", encoding="utf-8") as fh:
                    self.last_row = int(fh.read().strip() or self.last_row)
            except Exception:
                pass

    def _save_state(self) -> None:
        if not self.state_file:
            return
        try:
            with open(self.state_file, "w", encoding="utf-8") as fh:
                fh.write(str(self.last_row))
        except Exception:
            pass

    def _match(self, value: Any, matcher: Any) -> bool:
        if isinstance(matcher, dict):
            if "regex" in matcher:
                try:
                    return bool(re.search(matcher["regex"], str(value)))
                except re.error:
                    return False
            if "gt" in matcher:
                try:
                    return float(value) > float(matcher["gt"])
                except Exception:
                    return False
            if "gte" in matcher:
                try:
                    return float(value) >= float(matcher["gte"])
                except Exception:
                    return False
            if "lt" in matcher:
                try:
                    return float(value) < float(matcher["lt"])
                except Exception:
                    return False
            if "lte" in matcher:
                try:
                    return float(value) <= float(matcher["lte"])
                except Exception:
                    return False
        return value == matcher

    def poll(self) -> List[Dict[str, Any]]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pragma: no cover - dependency missing
            raise RuntimeError(
                "Excel triggers require the 'openpyxl' package. Install it with 'pip install openpyxl'."
            ) from exc

        file_path = self.config.get("file")
        sheet_name = self.config.get("sheet")
        filters: Dict[Any, Any] = self.config.get("filters", {})
        if not file_path:
            raise ValueError("file parameter required")

        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        max_row = getattr(ws, "max_row", None)
        if max_row is None:
            rows_attr = getattr(ws, "rows", [])
            try:
                max_row = len(list(rows_attr))
            except TypeError:
                max_row = len(rows_attr)

        results: List[Dict[str, Any]] = []
        for row_idx in range(self.last_row + 1, max_row + 1):
            cells = ws[row_idx]
            values = [getattr(c, "value", None) for c in cells]
            match = True
            for col, expected in filters.items():
                idx = int(col) - 1
                val = values[idx] if idx < len(values) else None
                if not self._match(val, expected):
                    match = False
                    break
            if match:
                results.append({"id": str(row_idx), "values": values})

        self.last_row = max_row
        self._save_state()
        return results


class ExcelCellChangeTrigger(BaseTrigger):
    """Trigger when monitored cells change value."""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.state_file = self.config.get("state_file")
        self.start_row = int(self.config.get("start_row", 2))
        self.columns = [int(c) for c in self.config.get("columns", [])]
        self._state: Dict[str, Any] = {}
        if self.state_file and os.path.exists(self.state_file):
            try:
                with open(self.state_file, "r", encoding="utf-8") as fh:
                    self._state = json.load(fh)
            except Exception:
                pass

    def _save_state(self) -> None:
        if not self.state_file:
            return
        try:
            with open(self.state_file, "w", encoding="utf-8") as fh:
                json.dump(self._state, fh)
        except Exception:
            pass

    def poll(self) -> List[Dict[str, Any]]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pragma: no cover - dependency missing
            raise RuntimeError(
                "Excel triggers require the 'openpyxl' package. Install it with 'pip install openpyxl'."
            ) from exc

        file_path = self.config.get("file")
        sheet_name = self.config.get("sheet")
        if not file_path:
            raise ValueError("file parameter required")

        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        max_row = getattr(ws, "max_row", None)
        if max_row is None:
            rows_attr = getattr(ws, "rows", [])
            try:
                max_row = len(list(rows_attr))
            except TypeError:
                max_row = len(rows_attr)

        changed: List[Dict[str, Any]] = []
        for row_idx in range(self.start_row, max_row + 1):
            cells = ws[row_idx]
            values = [getattr(c, "value", None) for c in cells]
            key = str(row_idx)
            prev = self._state.get(key, {})
            row_changed = False
            new_cols: Dict[int, Any] = {}
            for col in self.columns:
                idx = col - 1
                val = values[idx] if idx < len(values) else None
                if key in self._state and prev.get(str(col)) != val:
                    row_changed = True
                new_cols[str(col)] = val
            if row_changed:
                changed.append({"id": key, "values": values})
            self._state[key] = new_cols

        self._save_state()
        return changed


class ExcelFileUpdatedTrigger(BaseTrigger):
    """Trigger when the Excel file modification time changes."""

    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.state_file = self.config.get("state_file")
        self._last_mtime = 0.0
        if self.state_file and os.path.exists(self.state_file):
            try:
                with open(self.state_file, "r", encoding="utf-8") as fh:
                    self._last_mtime = float(fh.read().strip() or 0)
            except Exception:
                pass

    def _save_state(self) -> None:
        if not self.state_file:
            return
        try:
            with open(self.state_file, "w", encoding="utf-8") as fh:
                fh.write(str(self._last_mtime))
        except Exception:
            pass

    def poll(self) -> List[Dict[str, Any]]:
        file_path = self.config.get("file")
        if not file_path:
            raise ValueError("file parameter required")
        try:
            mtime = os.path.getmtime(file_path)
        except OSError:
            logging.error("Excel file not found: %s", file_path)
            return []

        if mtime != self._last_mtime:
            self._last_mtime = mtime
            self._save_state()
            return [{"id": str(mtime), "file": file_path}]
        return []


class ExcelAttachmentRowTrigger(ExcelRowAddedTrigger):
    """Trigger on new rows where an attachment column contains data."""

    def poll(self) -> List[Dict[str, Any]]:
        attachment_col = int(self.config.get("attachment_column", 0))
        rows = super().poll()
        results = []
        for row in rows:
            values = row.get("values", [])
            idx = attachment_col - 1
            if idx >= 0 and idx < len(values) and values[idx]:
                attachments = [s.strip() for s in str(values[idx]).split()]  # simple split
                row["attachments"] = attachments
                results.append(row)
        return results


class ExcelWriteRowAction(BaseAction):
    """Create or update rows in an Excel file."""

    def execute(self, data: Dict[str, Any]) -> None:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pragma: no cover - dependency missing
            raise RuntimeError(
                "Excel actions require the 'openpyxl' package. Install it with 'pip install openpyxl'."
            ) from exc

        file_path = self.params.get("file")
        sheet_name = self.params.get("sheet")
        row_idx = self.params.get("row")
        values = data.get("values") or list(data.values())
        if not file_path:
            raise ValueError("file parameter required")

        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        if row_idx:
            row_idx = int(row_idx)
            for col, val in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col, value=val)
        else:
            ws.append(values)
        wb.save(file_path)


class EmailSendAction(BaseAction):
    """Send an email using SMTP."""

    def execute(self, data: Dict[str, Any]) -> None:
        host = self.params.get("host", "localhost")
        port = int(self.params.get("port", 25))
        username = self.params.get("username")
        password = self.params.get("password")
        from_addr = self.params.get("from_addr")
        to_addr = self.params.get("to_addr") or data.get("to")
        subject = self.params.get("subject") or data.get("subject", "")
        body = self.params.get("body") or data.get("body", "")
        if not (from_addr and to_addr):
            raise ValueError("from_addr and to_addr required")

        import smtplib
        from email.message import EmailMessage

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = from_addr
        msg["To"] = to_addr
        msg.set_content(str(body))

        with smtplib.SMTP(host, port) as smtp:
            if username and password:
                smtp.login(username, password)
            smtp.send_message(msg)


class DBSaveAction(BaseAction):
    """Save data into a SQLite database."""

    def execute(self, data: Dict[str, Any]) -> None:
        import sqlite3

        db_path = self.params.get("db")
        table = self.params.get("table", "data")
        if not db_path:
            raise ValueError("db parameter required")
        fields = list(data.keys())
        placeholders = ",".join(["?" for _ in fields])
        cols = ",".join(fields)
        values = [data[f] for f in fields]
        conn = sqlite3.connect(db_path)
        with conn:
            conn.execute(
                f"CREATE TABLE IF NOT EXISTS {table} ({cols})"
            )
            conn.execute(
                f"INSERT INTO {table} ({cols}) VALUES ({placeholders})",
                values,
            )
        conn.close()


class FileCreateAction(BaseAction):
    """Create a file from row data."""

    def execute(self, data: Dict[str, Any]) -> None:
        path = self.params.get("path")
        fmt = self.params.get("format", "txt")
        if not path:
            raise ValueError("path parameter required")

        if fmt == "json":
            with open(path, "w", encoding="utf-8") as fh:
                json.dump(data, fh, ensure_ascii=False)
        elif fmt == "csv":
            import csv

            with open(path, "w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh)
                writer.writerow(list(data.keys()))
                writer.writerow(list(data.values()))
        else:
            with open(path, "w", encoding="utf-8") as fh:
                for k, v in data.items():
                    fh.write(f"{k}: {v}\n")


class AttachmentDownloadAction(BaseAction):
    """Download attachments referenced in a row."""

    def execute(self, data: Dict[str, Any]) -> Dict[str, Any]:
        urls = data.get("attachments", [])
        dest = self.params.get("dest")
        rename = self.params.get("rename")
        if not dest:
            raise ValueError("dest parameter required")
        os.makedirs(dest, exist_ok=True)
        downloaded: List[str] = []
        for url in urls:
            filename = os.path.basename(url)
            if rename:
                try:
                    filename = rename.format(**data, filename=filename)
                except Exception:
                    pass
            path = os.path.join(dest, filename)
            if url.startswith("http"):
                from urllib import request

                request.urlretrieve(url, path)
            else:
                try:
                    with open(url, "rb") as src, open(path, "wb") as dst:
                        dst.write(src.read())
                except Exception:
                    continue
            downloaded.append(path)
        data["downloaded"] = downloaded
        return data
