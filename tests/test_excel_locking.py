import sys
import types
import json
import threading
import time
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _setup_openpyxl(monkeypatch):
    openpyxl = types.ModuleType("openpyxl")

    class DummyCell:
        def __init__(self, value):
            self.value = value

    class Worksheet:
        def __init__(self):
            self.rows = []

        def append(self, values):
            self.rows.append(list(values))

        def __getitem__(self, idx):
            row = self.rows[idx - 1]
            return [DummyCell(v) for v in row]

        def cell(self, row, column, value=None):
            while len(self.rows) < row:
                self.rows.append([])
            r = self.rows[row - 1]
            while len(r) < column:
                r.append(None)
            r[column - 1] = value

    class Workbook:
        def __init__(self):
            self.active = Worksheet()

        def save(self, path):
            with open(path, "w", encoding="utf-8") as fh:
                json.dump(self.active.rows, fh)

        def __getitem__(self, name):
            return self.active

    def load_workbook(path, *args, **kwargs):
        wb = Workbook()
        if Path(path).exists():
            with open(path, "r", encoding="utf-8") as fh:
                wb.active.rows = json.load(fh)
        return wb

    openpyxl.Workbook = Workbook
    openpyxl.load_workbook = load_workbook
    monkeypatch.setitem(sys.modules, "openpyxl", openpyxl)


def test_write_row_waits_for_lock(monkeypatch, tmp_path):
    _setup_openpyxl(monkeypatch)
    import openpyxl
    from pyzap.plugins.excel_watch import ExcelWriteRowAction
    from pyzap.utils import excel_lock

    file_path = tmp_path / "book.xlsx"
    openpyxl.Workbook().save(file_path)

    action = ExcelWriteRowAction({"file": str(file_path)})

    def run():
        action.execute({"values": [1]})

    with excel_lock(file_path):
        t = threading.Thread(target=run)
        t.start()
        time.sleep(0.1)
        assert t.is_alive()
    t.join(1)
    assert not t.is_alive()
    wb = openpyxl.load_workbook(file_path)
    assert wb.active.rows == [[1]]


def test_append_waits_for_lock(monkeypatch, tmp_path):
    _setup_openpyxl(monkeypatch)
    import openpyxl
    from pyzap.plugins.excel_append import ExcelAppendAction
    from pyzap.utils import excel_lock

    file_path = tmp_path / "book.xlsx"
    openpyxl.Workbook().save(file_path)

    action = ExcelAppendAction({"file": str(file_path)})

    def run():
        action.execute({"values": [2]})

    with excel_lock(file_path):
        t = threading.Thread(target=run)
        t.start()
        time.sleep(0.1)
        assert t.is_alive()
    t.join(1)
    assert not t.is_alive()
    wb = openpyxl.load_workbook(file_path)
    assert wb.active.rows[-1] == [2]
