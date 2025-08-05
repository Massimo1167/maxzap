import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pyzap.plugins.excel_append import ExcelAppendAction


def _create_external_link_workbook(path: Path) -> None:
    """Create a minimal workbook containing an external link."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "='[book2.xlsx]Sheet1'!A1"
    wb.save(path)

    with zipfile.ZipFile(path, "r") as zin:
        files = {name: zin.read(name) for name in zin.namelist()}

    ns_ct = "http://schemas.openxmlformats.org/package/2006/content-types"
    ET.register_namespace("", ns_ct)
    ct_tree = ET.fromstring(files['[Content_Types].xml'])
    ET.SubElement(ct_tree, f"{{{ns_ct}}}Override", {
        "PartName": "/xl/externalLinks/externalLink1.xml",
        "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml",
    })
    files['[Content_Types].xml'] = ET.tostring(ct_tree, encoding="utf-8", xml_declaration=True)

    ns_rel = "http://schemas.openxmlformats.org/package/2006/relationships"
    ET.register_namespace("", ns_rel)
    rels_tree = ET.fromstring(files['xl/_rels/workbook.xml.rels'])
    ET.SubElement(rels_tree, f"{{{ns_rel}}}Relationship", {
        "Id": "rId2",
        "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink",
        "Target": "externalLinks/externalLink1.xml",
    })
    files['xl/_rels/workbook.xml.rels'] = ET.tostring(rels_tree, encoding="utf-8", xml_declaration=True)

    ns_wb = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ET.register_namespace("", ns_wb)
    ET.register_namespace("r", ns_r)
    wb_tree = ET.fromstring(files['xl/workbook.xml'])
    ext_refs = wb_tree.find(f"{{{ns_wb}}}externalReferences")
    if ext_refs is None:
        ext_refs = ET.SubElement(wb_tree, f"{{{ns_wb}}}externalReferences")
    ET.SubElement(ext_refs, f"{{{ns_wb}}}externalReference", {f"{{{ns_r}}}id": "rId2"})
    files['xl/workbook.xml'] = ET.tostring(wb_tree, encoding="utf-8", xml_declaration=True)

    ext_link_xml = (
        "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
        "<externalLink xmlns='http://schemas.openxmlformats.org/spreadsheetml/2006/main'>"
        "<externalBook><sheetNames><sheetName val='Sheet1'/></sheetNames></externalBook>"
        "</externalLink>"
    )
    files['xl/externalLinks/externalLink1.xml'] = ext_link_xml.encode("utf-8")

    rel_xml = (
        "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
        "<Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'>"
        "<Relationship Id='rId1' Type='http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath' Target='file:///book2.xlsx'/>"
        "</Relationships>"
    )
    files['xl/externalLinks/_rels/externalLink1.xml.rels'] = rel_xml.encode("utf-8")

    with zipfile.ZipFile(path, "w") as zout:
        for name, data in files.items():
            zout.writestr(name, data)


def test_excel_append_preserves_external_links(tmp_path):
    file_path = tmp_path / "book.xlsx"
    _create_external_link_workbook(file_path)

    action = ExcelAppendAction({"file": str(file_path)})
    action.execute({"a": 1})

    with zipfile.ZipFile(file_path) as z:
        assert "xl/externalLinks/externalLink1.xml" in z.namelist()

    wb = openpyxl.load_workbook(file_path, keep_links=True)
    assert len(getattr(wb, "_external_links", [])) == 1
